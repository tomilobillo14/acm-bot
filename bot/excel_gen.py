import os
import tempfile
from openpyxl import load_workbook
from coefs import calc_piso_coef, calc_sup_coef, calc_dias_coef

TEMPLATE = os.path.join(os.path.dirname(__file__), "TEMPLATE.xlsx")
COMP_ROWS = [(4,5), (6,7), (8,9), (10,11), (12,13)]

def _set(ws, row, col, value):
    try:
        ws.cell(row=row, column=col).value = value
    except AttributeError:
        pass  # celda fusionada no top-left

def generar_excel(tasado, comparables):
    wb = load_workbook(TEMPLATE)
    ws = wb["Modelo ACM con tiempo"]

    # Limpiar filas de comparables
    for dr, cr in COMP_ROWS:
        for col in range(1, 23):
            _set(ws, dr, col, None)
            _set(ws, cr, col, None)
        _set(ws, cr, 8,  0)
        _set(ws, cr, 16, 1.0)

    # Cargar comparables
    for i, (comp, (dr, cr)) in enumerate(zip(comparables, COMP_ROWS)):
        sup_cub  = comp.get("sup_cubierta", 0) or 0
        sup_semi = comp.get("sup_semi", 0) or 0
        sup_desc = comp.get("sup_desc", 0) or 0
        piso     = comp.get("piso", 0) or 0
        dias     = comp.get("dias_publicado", 0) or 0

        piso_c = calc_piso_coef(piso, True)
        sup_c  = calc_sup_coef(sup_cub)
        dias_c = calc_dias_coef(dias)

        _set(ws, dr,  1, i + 1)
        _set(ws, dr,  2, comp.get("direccion", "")[:50])
        _set(ws, dr,  3, "oferta")
        _set(ws, dr,  4, piso)
        _set(ws, dr,  5, comp.get("antiguedad", 0))
        _set(ws, dr,  6, f"{dias} Dias")
        _set(ws, dr,  7, comp.get("precio", 0))
        _set(ws, dr,  8, comp.get("cochera", 0))
        _set(ws, dr,  9, comp.get("sup_total", sup_cub))
        _set(ws, dr, 10, sup_cub)
        _set(ws, dr, 11, sup_semi)
        _set(ws, dr, 12, sup_desc)
        _set(ws, dr, 15, 1.0)   # Ub. — usuario ajusta en el Excel
        _set(ws, dr, 16, piso_c)
        _set(ws, dr, 17, sup_c)
        _set(ws, dr, 18, 1.0)   # Caract. Const. — default standard
        _set(ws, dr, 19, 1.0)   # Edad/Estado — default standard
        _set(ws, dr, 20, dias_c)
        _set(ws, cr,  8, 0)
        _set(ws, cr, 16, 1.0)

    # Fila amarilla del tasado (fila 11)
    yr         = 11
    sup_cub_t  = tasado.get("sup_cub",    0) or 0
    sup_semi_t = tasado.get("sup_semi",   0) or 0
    sup_desc_t = tasado.get("sup_desc",   0) or 0
    piso_t     = tasado.get("piso",       0) or 0
    asc_t      = tasado.get("ascensor",   True)
    dias_t     = tasado.get("dias",       0) or 0
    ub_t       = tasado.get("ub_coef",    1.0)
    estado_t   = tasado.get("estado_coef",1.0)

    piso_c_t = calc_piso_coef(piso_t, asc_t)
    sup_c_t  = calc_sup_coef(sup_cub_t)
    dias_c_t = calc_dias_coef(dias_t)

    _set(ws, yr,      1, tasado.get("direccion", "")[:50])
    _set(ws, yr,      4, piso_t)
    _set(ws, yr,      5, f"{tasado.get('antiguedad','?')} años")
    _set(ws, yr,      6, f"{dias_t} dias")
    _set(ws, yr,      8, tasado.get("cochera", 0))
    _set(ws, yr,      9, round(sup_cub_t + sup_semi_t + sup_desc_t, 2))
    _set(ws, yr,     10, sup_cub_t)
    _set(ws, yr,     11, sup_semi_t)
    _set(ws, yr,     12, sup_desc_t)
    _set(ws, yr,     15, ub_t)
    _set(ws, yr,     16, piso_c_t)
    _set(ws, yr,     17, sup_c_t)
    _set(ws, yr,     18, estado_t)
    _set(ws, yr,     19, estado_t)
    _set(ws, yr,     20, dias_c_t)
    _set(ws, yr + 1,  8, 0)
    _set(ws, yr + 1, 16, 1.0)

    tmp = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp)
    return tmp
